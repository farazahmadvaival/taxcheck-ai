import os
import csv
import openpyxl

def parse_excel_or_csv(file_path: str) -> tuple[str, str, float]:
    """
    Parses an Excel (.xlsx) or CSV (.csv) file deterministically.
    Returns: (extracted_text, extraction_method, confidence_score)
    """
    _, ext = os.path.splitext(file_path)
    ext = ext.lower()
    
    extracted_lines = []
    extraction_method = "DETERMINISTIC_EXCEL"
    confidence_score = 1.0  # Deterministic parsing has maximum confidence
    
    print(f"Parsing spreadsheet deterministically: {file_path}")
    
    if ext == ".csv":
        try:
            with open(file_path, mode="r", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                for row in reader:
                    # Skip completely empty rows
                    if any(row):
                        extracted_lines.append("\t".join(row))
        except Exception as e:
            print(f"Error parsing CSV {file_path}: {e}")
            raise e
            
    elif ext in [".xlsx", ".xlsm"]:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                extracted_lines.append(f"=== Sheet: {sheet_name} ===")
                for row in sheet.iter_rows(values_only=True):
                    # Convert all cell values to strings, mapping None to empty string
                    row_values = [str(cell) if cell is not None else "" for cell in row]
                    # Skip completely blank lines
                    if any(val.strip() for val in row_values):
                        extracted_lines.append("\t".join(row_values))
        except Exception as e:
            print(f"Error parsing Excel {file_path}: {e}")
            raise e
    else:
        raise ValueError(f"Unsupported spreadsheet format: {ext}")
        
    extracted_text = "\n".join(extracted_lines)
    return extracted_text, extraction_method, confidence_score
